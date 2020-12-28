import numpy as np
import tkinter as tk
import openpyxl as xl

from tkinter import filedialog as fd
from tkinter.messagebox import showerror


class Main(tk.Frame):
    def __init__(self, root):
        super().__init__(root)

        self.root = root
        self.create_window()

        self.experts = []
        self.objects = []
        self.expert_data_list_1 = []
        self.expert_data_list_2 = []
        self.expert_data_list_3 = []

        self.objects_added = False

        self.expert_dict = dict()
        self.weight = dict()
        self.terms = ['Приспособленность языка',
                      'Скорость разработки', 'Сложность разработки']

        self.experts_entries = []

    def create_window(self):
        # top frame
        top_frame = tk.Frame(self)
        top_frame.pack(side=tk.TOP)

        tk.Label(top_frame, text='Эксперт: ', font='Times 14',
                 fg='#000080').pack(side=tk.LEFT, padx=5, pady=5)

        self.expert_entry = tk.Entry(top_frame, font='Times 14', fg='#000080')
        self.expert_entry.pack(side=tk.LEFT, padx=5, pady=5)

        tk.Button(top_frame, text='Добавить данные эксперта', command=self.on_add_expert_data_btn_clicked,
                  font='Times 14', fg='#000080').pack(side=tk.LEFT, padx=5, pady=5)

        self.experts_lbl = tk.Label(
            top_frame, text='', font='Times 14', fg='#000080')
        self.experts_lbl.pack(side=tk.LEFT, padx=5, pady=5)

        # center frame
        center_frame = tk.Frame(self)
        center_frame.pack(side=tk.TOP)
        tk.Button(center_frame, text='Ввести веса экспертам', command=self.on_compare_experts_btn_clicked,
                  font='Times 14', fg='#000080').pack(side=tk.LEFT, padx=5, pady=5)

        # last frame
        frame = tk.Frame(self)
        frame.pack(side=tk.TOP)

        tk.Button(frame, text='Посчитать результат', command=self.calc_result,
                  font='Times 14', fg='#000080').pack(side=tk.LEFT, padx=5, pady=5)

        frame_1 = tk.Frame(self)
        frame_1.pack(side=tk.TOP)

        self.result_text = tk.Text(
            frame_1, width=100, height=25, font='Times 14', fg='#000080')
        self.result_text.pack(side=tk.TOP, padx=5, pady=5)

    def on_add_expert_data_btn_clicked(self):
        text = self.expert_entry.get()
        if text != '' and text not in self.experts:
            filename = fd.askopenfilename()
            wb = xl.load_workbook(filename=filename)
            sheet = wb['Приспособленность языка']

            if not self.objects_added:
                row = 2
                while sheet.cell(row=row, column=1).value != None:
                    print(str(sheet.cell(row=row, column=1).value))
                    self.objects.append(
                        str(sheet.cell(row=row, column=1).value))
                    row += 1
                self.objects_added = True

            length = len(self.objects)

            matrix = np.zeros((length, length), dtype=np.float)

            for i in range(2, len(self.objects) + 2):
                for j in range(2, len(self.objects) + 2):
                    val = str(sheet.cell(row=i, column=j).value)
                    if val == '1/2':
                        matrix[i-2][j-2] = 1 / 2
                    elif val == '1/3':
                        matrix[i-2][j-2] = 1 / 3
                    elif val == '1/4':
                        matrix[i-2][j-2] = 1 / 4
                    elif val == '1/5':
                        matrix[i-2][j-2] = 1 / 5
                    elif val == '1/6':
                        matrix[i-2][j-2] = 1 / 6
                    elif val == '1/7':
                        matrix[i-2][j-2] = 1 / 7
                    elif val == '1/8':
                        matrix[i-2][j-2] = 1 / 8
                    elif val == '1/9':
                        matrix[i-2][j-2] = 1 / 9
                    elif float(val) > 0 and float(val) < 10:
                        matrix[i-2][j-2] = float(val)
                    else:
                        showerror(title="Error", message="Некорректные данные")
                        return
            self.expert_data_list_1.append(matrix)

            sheet = wb['Скорость разработки']
            matrix_1 = np.zeros((length, length), dtype=np.float)

            for i in range(2, len(self.objects) + 2):
                for j in range(2, len(self.objects) + 2):
                    val = str(sheet.cell(row=i, column=j).value)
                    if val == '1/2':
                        matrix_1[i-2][j-2] = 1 / 2
                    elif val == '1/3':
                        matrix_1[i-2][j-2] = 1 / 3
                    elif val == '1/4':
                        matrix_1[i-2][j-2] = 1 / 4
                    elif val == '1/5':
                        matrix_1[i-2][j-2] = 1 / 5
                    elif val == '1/6':
                        matrix_1[i-2][j-2] = 1 / 6
                    elif val == '1/7':
                        matrix_1[i-2][j-2] = 1 / 7
                    elif val == '1/8':
                        matrix_1[i-2][j-2] = 1 / 8
                    elif val == '1/9':
                        matrix_1[i-2][j-2] = 1 / 9
                    elif float(val) > 0 and float(val) < 10:
                        matrix_1[i-2][j-2] = float(val)
                    else:
                        showerror(title="Error", message="Некорректные данные")
                        return
            self.expert_data_list_2.append(matrix_1)

            sheet = wb['Сложность разработки']
            matrix_2 = np.zeros((length, length), dtype=np.float)

            for i in range(2, len(self.objects) + 2):
                for j in range(2, len(self.objects) + 2):
                    val = str(sheet.cell(row=i, column=j).value)
                    if val == '1/2':
                        matrix_2[i-2][j-2] = 1 / 2
                    elif val == '1/3':
                        matrix_2[i-2][j-2] = 1 / 3
                    elif val == '1/4':
                        matrix_2[i-2][j-2] = 1 / 4
                    elif val == '1/5':
                        matrix_2[i-2][j-2] = 1 / 5
                    elif val == '1/6':
                        matrix_2[i-2][j-2] = 1 / 6
                    elif val == '1/7':
                        matrix_2[i-2][j-2] = 1 / 7
                    elif val == '1/8':
                        matrix_2[i-2][j-2] = 1 / 8
                    elif val == '1/9':
                        matrix_2[i-2][j-2] = 1 / 9
                    elif float(val) > 0 and float(val) < 10:
                        matrix_2[i-2][j-2] = float(val)
                    else:
                        showerror(title="Error", message="Некорректные данные")
                        return
            self.expert_data_list_3.append(matrix_2)

            data_1 = self._check_expert(matrix)
            data_2 = self._check_expert(matrix_1)
            data_3 = self._check_expert(matrix_2)

            if data_1 == None:
                showerror(
                    title='Error', message='Эксперт противорчит себе в данных \"Приспособленность языка\"!')
                return
            elif data_2 == None:
                showerror(
                    title='Error', message='Эксперт противорчит себе в данных \"Скорость разработки\"!')
                return
            elif data_3 == None:
                showerror(
                    title='Error', message='Эксперт противорчит себе в данных \"Сложность разработки\"!')
                return
            else:
                self._add_values(data_1, data_2, data_3, text)
                self.experts.append(text)

                exp = 'Выбранные эксперты: ' + \
                    str(self.experts).replace(
                        '[', '').replace(']', '').replace('\'', '')
                self.experts_lbl['text'] = exp

            # print(self.expert_data_list_1)
            # print(self.expert_data_list_2)
            # print(self.expert_data_list_3)
        else:
            showerror(title='Error', message='Такой эксперт уже задан')

    # проверка эксперта
    def _check_expert(self, matrix):
        n = len(self.objects)
        arr = np.array(matrix)
        a, v = np.linalg.eig(arr)
        lbd = max(a)

        print('a = ', a)
        print('v = ', v)

        data = []

        for i in range(0, n):
            value = matrix[i]
            data.append(value.prod()**(1.0/len(value)))

        k = (lbd - n) / (n - 1)
        print(k)
        if k < 0.3:
            norm = max(data)
            for i in range(0, len(data)):
                data[i] /= norm

            return data
        else:
            return None

    # добавление значений
    def _add_values(self, data_1, data_2, data_3, expert):
        values = dict()
        values[self.terms[0]] = data_1
        values[self.terms[1]] = data_2
        values[self.terms[2]] = data_3
        self.expert_dict[expert] = values
        print(self.expert_dict)

    def open_window(self, _class):
        try:
            if self.new.state() == 'normal':
                self.new.focus()
        except:
            self.new = tk.Toplevel(self.root)
            _class(self.new, self)

    def on_compare_experts_btn_clicked(self):
        self.open_window(CompareExpertsWindow)

    def set_weight(self, weight):
        self.weight = weight
        print(self.weight)

    def calc_result(self):
        result = dict()
        experts = self.experts.copy()
        weight = self.weight.copy()

        while(True):
            print('\n--------------------------------\n')
            print('exp = ', experts)
            print('weight = ', weight)
            flag = True
            for term in self.terms:
                temp_result = list()
                for obj in range(0, len(self.objects)):
                    rating = list()
                    for exp in experts:
                        r = self.expert_dict[exp][term]
                        rating.append(r[obj] * weight[exp])

                    print('rating = ', rating)

                    if not(self.coef_variation(rating)):
                        flag = False
                        break

                    temp_result.append(sum(rating) / len(rating))
                print('temp_RESULT', temp_result)
                if not(flag):
                    if len(experts) > 1:
                        flag = True
                        _min = list(weight.keys())
                        _min = _min[0]
                        for k, v in weight.items():
                            if v < weight[_min]:
                                _min = k
                        weight.pop(_min)
                        experts.remove(_min)
                        # result[term] = temp_result
                        # break
                    else:
                        print("123123123123")
                        showerror(
                            title='Error', message='Не удалось достичь согласованности экспертов!')
                        break
                # else:
                #     print('44444444444')
                #     showerror(title='Error', message='Не удалось достичь согласованности экспертов!')
                #     break
                print('temp_result', temp_result)
                result[term] = temp_result
            print("AOOOOOOOOOOOOOOOOOOOOOOOOOo")
            flag = True
            for term in self.terms:
                if result[term] == []:
                    flag = False
            if flag:
                break

        print('#####################\nresult = ', result)
        self.print_result(result)

    def print_result(self, result):
        self.result_text.delete('1.0', 'end')
        for k, v in result.items():
            self.result_text.insert('end-1c', k + ':\n')
            for i in range(0, len(v)):
                self.result_text.insert(
                    'end-1c', '\t' + self.objects[i] + ':' + str(v[i])[:4] + ' ')
            self.result_text.insert('end-1c', '\n')

    def coef_variation(self, _list):
        arr = np.array(_list)
        print('arr = ', arr)
        sigma = np.std(arr)
        sr = sum(_list) / len(_list)
        print('sigma = ', sigma)
        print('sr = ', sr)
        print('k = ', sigma / sr)
        return sigma / sr <= 0.4


class CompareExpertsWindow(tk.Frame):

    def __init__(self, root, app):
        super().__init__(root)

        self.app = app
        self.root = root
        self.root.geometry("600x300+200+200")
        self.weight = dict()
        self.experts = self.app.experts
        self.experts_entries = []
        for i in self.experts:
            self.weight[i] = dict()
        self.create_window()

    def create_window(self):
        self.master.title('Сравнить экспертов')
        self.pack(fill=tk.BOTH, expand=True)
        self.table_frame = tk.Frame(self)
        self.table_frame.pack(side=tk.TOP, padx=5, pady=5)

        n = len(self.experts)
        for i in range(0, n + 1):
            for j in range(1, n + 1):
                if i == 0 and j != 0:
                    tk.Label(
                        self.table_frame, text=self.experts[j-1], font='Times 14', fg='#000080').grid(row=i, column=j)
                    tk.Label(
                        self.table_frame, text=self.experts[j-1], font='Times 14', fg='#000080').grid(row=j, column=i)
                else:
                    entry = tk.Entry(self.table_frame,
                                     font='Times 14', fg='#000080')
                    entry.grid(row=i, column=j)
                    self.experts_entries.append(entry)

        frame_2 = tk.Frame(self)
        frame_2.pack(side=tk.TOP, expand=True, padx=5, pady=5, fill=tk.X)

        ok_btn = tk.Button(
            frame_2, text='OK', command=self.check_weights, font='Times 14', fg='#000080')
        ok_btn.pack(fill=tk.X, padx=5)

    def check_weights(self):
        data = list()
        n = len(self.experts)
        matrix = list()

        values = list()
        for item in self.experts_entries:
            value = item.get()

            if value == "":
                values.append(1.0)
            elif value == '1/2':
                values.append(1/2)
            elif value == '1/3':
                values.append(1/3)
            elif value == '1/4':
                values.append(1/4)
            elif value == '1/5':
                values.append(1/5)
            elif value == '1/6':
                values.append(1/6)
            elif value == '1/7':
                values.append(1/7)
            elif value == '1/8':
                values.append(1/8)
            elif value == '1/9':
                values.append(1/9)
            elif int(value) > 0 and int(value) < 10:
                values.append(float(value))
            else:
                showerror(title='Error', message='Некорректные данные!')

        for i in range(0, n):
            value = np.array(values[i*n:(i+1)*n])
            matrix.append(value)
            data.append(value.prod()**(1.0/len(value)))

        matrix = np.array(matrix)
        a, v = np.linalg.eig(matrix)
        lbd = max(a)

        if (lbd - n) / (n - 1) < 0.1:
            norm = sum(data)
            for i in range(0, len(data)):
                data[i] /= norm
            self.add_values(data)
        else:
            showerror(title='Error', message='Данные противоречевы!')
            return

        self.on_ok_btn_clicked()

    def add_values(self, data):
        for i in range(0, len(self.experts)):
            self.weight[self.experts[i]] = data[i]

    def on_ok_btn_clicked(self):
        self.app.set_weight(self.weight)
        self.root.destroy()


if __name__ == '__main__':
    root = tk.Tk()
    app = Main(root)
    app.grid()
    root.title("Лаб 3")
    root.geometry("900x700")
    root.mainloop()
